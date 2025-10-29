#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Planillas Multirégimen - Perú
Versión Completa con Todas las Funcionalidades
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
import os
import sqlite3
import io
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

# Configuración de Flask
app = Flask(__name__)
app.config['SECRET_KEY'] = 'planillas_peru_2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///planillas_peru.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'

# Crear directorio de uploads
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Inicializar SQLAlchemy
db = SQLAlchemy(app)

# Modelos completos
class Empresa(db.Model):
    __tablename__ = 'empresas'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(200), nullable=False)
    ruc = db.Column(db.String(11), unique=True, nullable=False)
    regimen_laboral = db.Column(db.String(50), nullable=False)
    direccion = db.Column(db.String(500))
    telefono = db.Column(db.String(20))
    email = db.Column(db.String(100))
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

class Empleado(db.Model):
    __tablename__ = 'empleados'
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresas.id'), nullable=False)
    nombres = db.Column(db.String(100), nullable=False)
    apellidos = db.Column(db.String(100), nullable=False)
    dni = db.Column(db.String(8), nullable=False)
    sueldo_base = db.Column(db.Float, nullable=False)
    fecha_ingreso = db.Column(db.Date, nullable=False)
    fecha_nacimiento = db.Column(db.Date)
    direccion = db.Column(db.String(500))
    telefono = db.Column(db.String(20))
    email = db.Column(db.String(100))
    tipo_pension = db.Column(db.String(20), default='ONP')
    afp_codigo = db.Column(db.String(10))
    cuenta_bancaria = db.Column(db.String(20))
    banco = db.Column(db.String(50))
    tipo_pago = db.Column(db.String(20), default='mensual')  # 'mensual' o 'quincenal'
    descuento_alimentos = db.Column(db.Float, default=0.0)  # Descuento judicial por alimentos
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

class Locador(db.Model):
    __tablename__ = 'locadores'
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresas.id'), nullable=False)
    nombres = db.Column(db.String(100), nullable=False)
    apellidos = db.Column(db.String(100), nullable=False)
    dni = db.Column(db.String(8), nullable=False)
    monto_mensual = db.Column(db.Float, nullable=False)
    fecha_inicio = db.Column(db.Date, nullable=False)
    suspendido = db.Column(db.Boolean, default=False)
    cuenta_bancaria = db.Column(db.String(20))
    banco = db.Column(db.String(50))
    descuento_alimentos = db.Column(db.Float, default=0.0)  # Descuento judicial por alimentos
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

class Ausencia(db.Model):
    __tablename__ = 'ausencias'
    id = db.Column(db.Integer, primary_key=True)
    empleado_id = db.Column(db.Integer, db.ForeignKey('empleados.id'), nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    tipo = db.Column(db.String(20), nullable=False)  # 'justificada', 'injustificada', 'permiso'
    motivo = db.Column(db.String(500))
    horas_perdidas = db.Column(db.Float, default=0.0)
    empleado = db.relationship('Empleado', backref='ausencias')

class Prestamo(db.Model):
    __tablename__ = 'prestamos'
    id = db.Column(db.Integer, primary_key=True)
    empleado_id = db.Column(db.Integer, db.ForeignKey('empleados.id'), nullable=False)
    monto = db.Column(db.Float, nullable=False)
    cuotas = db.Column(db.Integer, nullable=False)
    cuota_mensual = db.Column(db.Float, nullable=False)
    cuotas_pagadas = db.Column(db.Integer, default=0)
    fecha_inicio = db.Column(db.Date, nullable=False)
    activo = db.Column(db.Boolean, default=True)
    empleado = db.relationship('Empleado', backref='prestamos')

class Adelanto(db.Model):
    __tablename__ = 'adelantos'
    id = db.Column(db.Integer, primary_key=True)
    empleado_id = db.Column(db.Integer, db.ForeignKey('empleados.id'), nullable=False)
    monto = db.Column(db.Float, nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    descuento_mensual = db.Column(db.Float, nullable=False)
    meses_restantes = db.Column(db.Integer, nullable=False)
    activo = db.Column(db.Boolean, default=True)
    empleado = db.relationship('Empleado', backref='adelantos')

# Rutas principales
@app.route('/')
def index():
    """Página principal con resumen de empresas"""
    empresas = Empresa.query.all()
    empresas_con_conteos = []
    
    for empresa in empresas:
        # Contar empleados y locadores activos
        empleados_activos = Empleado.query.filter_by(empresa_id=empresa.id, activo=True).count()
        locadores_activos = Locador.query.filter_by(empresa_id=empresa.id, activo=True).count()
        
        empresas_con_conteos.append({
            'empresa': empresa,
            'empleados_activos': empleados_activos,
            'locadores_activos': locadores_activos
        })
    
    return render_template('index.html', empresas_con_conteos=empresas_con_conteos)

@app.route('/empresas')
def empresas():
    """Lista de empresas"""
    empresas = Empresa.query.all()
    return render_template('empresas.html', empresas=empresas)

@app.route('/empresa/nueva', methods=['GET', 'POST'])
def nueva_empresa():
    """Crear nueva empresa"""
    if request.method == 'POST':
        try:
            empresa = Empresa(
                nombre=request.form['nombre'],
                ruc=request.form['ruc'],
                regimen_laboral=request.form['regimen_laboral'],
                direccion=request.form.get('direccion', ''),
                telefono=request.form.get('telefono', ''),
                email=request.form.get('email', '')
            )
            db.session.add(empresa)
            db.session.commit()
            flash('Empresa creada exitosamente', 'success')
            return redirect(url_for('empresas'))
        except Exception as e:
            flash(f'Error al crear empresa: {str(e)}', 'error')
    
    return render_template('nueva_empresa.html')

@app.route('/personal/<int:empresa_id>')
def personal(empresa_id):
    """Lista de personal de una empresa"""
    empresa = Empresa.query.get_or_404(empresa_id)
    empleados = Empleado.query.filter_by(empresa_id=empresa_id, activo=True).all()
    locadores = Locador.query.filter_by(empresa_id=empresa_id, activo=True).all()
    return render_template('personal.html', empresa=empresa, empleados=empleados, locadores=locadores)

@app.route('/empleado/nuevo/<int:empresa_id>', methods=['GET', 'POST'])
def nuevo_empleado(empresa_id):
    """Crear nuevo empleado"""
    empresa = Empresa.query.get_or_404(empresa_id)
    
    if request.method == 'POST':
        try:
            empleado = Empleado(
                empresa_id=empresa_id,
                nombres=request.form['nombres'],
                apellidos=request.form['apellidos'],
                dni=request.form['dni'],
                sueldo_base=float(request.form['sueldo_base']),
                fecha_ingreso=datetime.strptime(request.form['fecha_ingreso'], '%Y-%m-%d').date(),
                fecha_nacimiento=datetime.strptime(request.form['fecha_nacimiento'], '%Y-%m-%d').date() if request.form.get('fecha_nacimiento') else None,
                direccion=request.form.get('direccion', ''),
                telefono=request.form.get('telefono', ''),
                email=request.form.get('email', ''),
                tipo_pension=request.form.get('tipo_pension', 'ONP'),
                afp_codigo=request.form.get('afp_codigo', '') if request.form.get('tipo_pension') == 'AFP' else None,
                cuenta_bancaria=request.form.get('cuenta_bancaria', ''),
                banco=request.form.get('banco', ''),
                tipo_pago=request.form.get('tipo_pago', 'mensual'),
                descuento_alimentos=float(request.form.get('descuento_alimentos', 0))
            )
            db.session.add(empleado)
            db.session.commit()
            flash('Empleado creado exitosamente', 'success')
            return redirect(url_for('personal', empresa_id=empresa_id))
        except Exception as e:
            flash(f'Error al crear empleado: {str(e)}', 'error')
    
    return render_template('nuevo_empleado_completo.html', empresa=empresa)

@app.route('/locador/nuevo/<int:empresa_id>', methods=['GET', 'POST'])
def nuevo_locador(empresa_id):
    """Crear nuevo locador"""
    empresa = Empresa.query.get_or_404(empresa_id)
    
    if request.method == 'POST':
        try:
            locador = Locador(
                empresa_id=empresa_id,
                nombres=request.form['nombres'],
                apellidos=request.form['apellidos'],
                dni=request.form['dni'],
                monto_mensual=float(request.form['monto_mensual']),
                fecha_inicio=datetime.strptime(request.form['fecha_inicio'], '%Y-%m-%d').date(),
                suspendido=request.form.get('suspendido') == 'on',
                cuenta_bancaria=request.form.get('cuenta_bancaria', ''),
                banco=request.form.get('banco', ''),
                descuento_alimentos=float(request.form.get('descuento_alimentos', 0))
            )
            db.session.add(locador)
            db.session.commit()
            flash('Locador creado exitosamente', 'success')
            return redirect(url_for('personal', empresa_id=empresa_id))
        except Exception as e:
            flash(f'Error al crear locador: {str(e)}', 'error')
    
    return render_template('nuevo_locador_completo.html', empresa=empresa)

@app.route('/empleado/editar/<int:empleado_id>', methods=['GET', 'POST'])
def editar_empleado(empleado_id):
    """Editar empleado existente"""
    empleado = Empleado.query.get_or_404(empleado_id)
    empresa = Empresa.query.get(empleado.empresa_id)
    
    if request.method == 'POST':
        try:
            empleado.nombres = request.form['nombres']
            empleado.apellidos = request.form['apellidos']
            empleado.dni = request.form['dni']
            empleado.sueldo_base = float(request.form['sueldo_base'])
            empleado.fecha_ingreso = datetime.strptime(request.form['fecha_ingreso'], '%Y-%m-%d').date()
            empleado.fecha_nacimiento = datetime.strptime(request.form['fecha_nacimiento'], '%Y-%m-%d').date() if request.form.get('fecha_nacimiento') else None
            empleado.direccion = request.form.get('direccion', '')
            empleado.telefono = request.form.get('telefono', '')
            empleado.email = request.form.get('email', '')
            empleado.tipo_pension = request.form.get('tipo_pension', 'ONP')
            empleado.afp_codigo = request.form.get('afp_codigo', '') if request.form.get('tipo_pension') == 'AFP' else None
            empleado.cuenta_bancaria = request.form.get('cuenta_bancaria', '')
            empleado.banco = request.form.get('banco', '')
            empleado.tipo_pago = request.form.get('tipo_pago', 'mensual')
            empleado.descuento_alimentos = float(request.form.get('descuento_alimentos', 0))
            
            db.session.commit()
            flash('Empleado actualizado exitosamente', 'success')
            return redirect(url_for('personal', empresa_id=empleado.empresa_id))
        except Exception as e:
            flash(f'Error al actualizar empleado: {str(e)}', 'error')
    
    return render_template('editar_empleado.html', empleado=empleado, empresa=empresa)

@app.route('/locador/editar/<int:locador_id>', methods=['GET', 'POST'])
def editar_locador(locador_id):
    """Editar locador existente"""
    locador = Locador.query.get_or_404(locador_id)
    empresa = Empresa.query.get(locador.empresa_id)
    
    if request.method == 'POST':
        try:
            locador.nombres = request.form['nombres']
            locador.apellidos = request.form['apellidos']
            locador.dni = request.form['dni']
            locador.monto_mensual = float(request.form['monto_mensual'])
            locador.fecha_inicio = datetime.strptime(request.form['fecha_inicio'], '%Y-%m-%d').date()
            locador.suspendido = request.form.get('suspendido') == 'on'
            locador.cuenta_bancaria = request.form.get('cuenta_bancaria', '')
            locador.banco = request.form.get('banco', '')
            locador.descuento_alimentos = float(request.form.get('descuento_alimentos', 0))
            
            db.session.commit()
            flash('Locador actualizado exitosamente', 'success')
            return redirect(url_for('personal', empresa_id=locador.empresa_id))
        except Exception as e:
            flash(f'Error al actualizar locador: {str(e)}', 'error')
    
    return render_template('editar_locador.html', locador=locador, empresa=empresa)

@app.route('/empleado/eliminar/<int:empleado_id>', methods=['POST'])
def eliminar_empleado(empleado_id):
    """Eliminar empleado (marcar como inactivo)"""
    empleado = Empleado.query.get_or_404(empleado_id)
    empleado.activo = False
    db.session.commit()
    flash('Empleado eliminado exitosamente', 'success')
    return redirect(url_for('personal', empresa_id=empleado.empresa_id))

@app.route('/locador/eliminar/<int:locador_id>', methods=['POST'])
def eliminar_locador(locador_id):
    """Eliminar locador (marcar como inactivo)"""
    locador = Locador.query.get_or_404(locador_id)
    locador.activo = False
    db.session.commit()
    flash('Locador eliminado exitosamente', 'success')
    return redirect(url_for('personal', empresa_id=locador.empresa_id))

@app.route('/planilla/<int:empresa_id>')
def planilla(empresa_id):
    """Cálculo de planilla"""
    empresa = Empresa.query.get_or_404(empresa_id)
    empleados = Empleado.query.filter_by(empresa_id=empresa_id, activo=True).all()
    locadores = Locador.query.filter_by(empresa_id=empresa_id, activo=True).all()
    
    return render_template('planilla_completo.html', empresa=empresa, empleados=empleados, locadores=locadores)

@app.route('/calcular_planilla/<int:empresa_id>', methods=['POST'])
def calcular_planilla(empresa_id):
    """Calcular planilla para un mes específico"""
    empresa = Empresa.query.get_or_404(empresa_id)
    mes = int(request.form['mes'])
    año = int(request.form['año'])
    
    # Obtener empleados y locadores activos
    empleados = Empleado.query.filter_by(empresa_id=empresa_id, activo=True).all()
    locadores = Locador.query.filter_by(empresa_id=empresa_id, activo=True).all()
    
    # Calcular planilla
    resultado = calcular_planilla_completa(empresa_id, mes, año)
    
    return render_template('planilla_completo.html', 
                         empresa=empresa, 
                         empleados=empleados, 
                         locadores=locadores,
                         resultado=resultado,
                         mes=mes,
                         año=año)

@app.route('/exportar_excel/<int:empresa_id>', methods=['POST'])
def exportar_excel(empresa_id):
    """Exportar planilla a Excel"""
    empresa = Empresa.query.get_or_404(empresa_id)
    mes = int(request.form['mes'])
    año = int(request.form['año'])
    
    try:
        resultado = calcular_planilla_completa(empresa_id, mes, año)
        
        # Crear libro de trabajo
        wb = Workbook()
        
        # Hoja de empleados
        ws_empleados = wb.active
        ws_empleados.title = "Empleados"
        
        # Encabezados para empleados
        headers_empleados = [
            'Nombres', 'Apellidos', 'DNI', 'Sueldo Base', 'Sueldo Ajustado',
            'Vacaciones', 'CTS', 'Gratificación', 'Asignación Familiar',
            'Pensión', 'Impuesto Renta', 'Descuento Alimentos', 'Préstamos', 'Adelantos',
            'Total Ingresos', 'Total Descuentos', 'Neto a Pagar'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers_empleados, 1):
            ws_empleados.cell(row=1, column=col, value=header)
        
        # Escribir datos de empleados
        for row, emp in enumerate(resultado['empleados'], 2):
            ws_empleados.cell(row=row, column=1, value=emp['empleado'].nombres)
            ws_empleados.cell(row=row, column=2, value=emp['empleado'].apellidos)
            ws_empleados.cell(row=row, column=3, value=emp['empleado'].dni)
            ws_empleados.cell(row=row, column=4, value=emp['sueldo_ajustado'])
            ws_empleados.cell(row=row, column=5, value=emp['sueldo_ajustado'])
            ws_empleados.cell(row=row, column=6, value=emp['vacaciones'])
            ws_empleados.cell(row=row, column=7, value=emp['cts'])
            ws_empleados.cell(row=row, column=8, value=emp['gratificacion'])
            ws_empleados.cell(row=row, column=9, value=emp['asignacion_familiar'])
            ws_empleados.cell(row=row, column=10, value=emp['pension'])
            ws_empleados.cell(row=row, column=11, value=emp['impuesto_renta'])
            ws_empleados.cell(row=row, column=12, value=emp['descuento_alimentos'])
            ws_empleados.cell(row=row, column=13, value=emp['descuentos']['prestamos'])
            ws_empleados.cell(row=row, column=14, value=emp['descuentos']['adelantos'])
            ws_empleados.cell(row=row, column=15, value=emp['total_ingresos'])
            ws_empleados.cell(row=row, column=16, value=emp['total_descuentos'])
            ws_empleados.cell(row=row, column=17, value=emp['neto'])
        
        # Hoja de locadores
        ws_locadores = wb.create_sheet("Locadores")
        
        # Encabezados para locadores
        headers_locadores = [
            'Nombres', 'Apellidos', 'DNI', 'Monto Base', 'Retención 4ta Cat.',
            'Descuento Alimentos', 'Total Ingresos', 'Total Descuentos', 'Neto a Pagar'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers_locadores, 1):
            ws_locadores.cell(row=1, column=col, value=header)
        
        # Escribir datos de locadores
        for row, loc in enumerate(resultado['locadores'], 2):
            ws_locadores.cell(row=row, column=1, value=loc['locador'].nombres)
            ws_locadores.cell(row=row, column=2, value=loc['locador'].apellidos)
            ws_locadores.cell(row=row, column=3, value=loc['locador'].dni)
            ws_locadores.cell(row=row, column=4, value=loc['monto_base'])
            ws_locadores.cell(row=row, column=5, value=loc['retencion_4ta'])
            ws_locadores.cell(row=row, column=6, value=loc['descuento_alimentos'])
            ws_locadores.cell(row=row, column=7, value=loc['total_ingresos'])
            ws_locadores.cell(row=row, column=8, value=loc['total_descuentos'])
            ws_locadores.cell(row=row, column=9, value=loc['neto'])
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=planilla_{empresa.nombre}_{mes:02d}_{año}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Error al exportar: {str(e)}', 'error')
        return redirect(url_for('planilla', empresa_id=empresa_id))

def calcular_planilla_completa(empresa_id, mes, año):
    """Función completa de cálculo de planilla con todas las funcionalidades"""
    empresa = Empresa.query.get(empresa_id)
    empleados = Empleado.query.filter_by(empresa_id=empresa_id, activo=True).all()
    locadores = Locador.query.filter_by(empresa_id=empresa_id, activo=True).all()
    
    resultado = {
        'empresa': empresa,
        'mes': mes,
        'año': año,
        'empleados': [],
        'locadores': [],
        'totales': {
            'total_ingresos': 0,
            'total_descuentos': 0,
            'total_neto': 0
        }
    }
    
    # Calcular para empleados
    for empleado in empleados:
        sueldo_base = empleado.sueldo_base
        
        # Aplicar pago quincenal si corresponde
        sueldo_ajustado = sueldo_base
        if empleado.tipo_pago == 'quincenal':
            sueldo_ajustado = sueldo_base * 0.5
        
        # Calcular beneficios según régimen
        if empresa.regimen_laboral == 'microempresa':
            vacaciones = 0
            cts = 0
            gratificacion = 0
            asignacion_familiar = 0
        elif empresa.regimen_laboral == 'pequeña_empresa':
            vacaciones = sueldo_ajustado * 15 / 365
            cts = sueldo_ajustado * 15 / 365
            gratificacion = sueldo_ajustado * 0.5 if mes in [7, 12] else 0
            asignacion_familiar = 102.50 if sueldo_ajustado <= 1025 else 0
        else:  # régimen general
            vacaciones = sueldo_ajustado * 30 / 365
            cts = sueldo_ajustado
            gratificacion = sueldo_ajustado * 1.09 if mes in [7, 12] else 0
            asignacion_familiar = 102.50 if sueldo_ajustado <= 1025 else 0
        
        # Calcular descuentos
        pension = sueldo_ajustado * 0.13 if empleado.tipo_pension == 'ONP' else sueldo_ajustado * 0.12
        impuesto_renta = max(0, (sueldo_ajustado - 1025) * 0.08) if sueldo_ajustado > 1025 else 0
        
        # Descuento por alimentos (solo en pago mensual)
        descuento_alimentos = empleado.descuento_alimentos if empleado.tipo_pago == 'mensual' else 0
        
        # Calcular préstamos y adelantos
        prestamos = sum(p.cuota_mensual for p in empleado.prestamos if p.activo)
        adelantos = sum(a.descuento_mensual for a in empleado.adelantos if a.activo)
        
        total_ingresos = sueldo_ajustado + vacaciones + cts + gratificacion + asignacion_familiar
        total_descuentos = pension + impuesto_renta + descuento_alimentos + prestamos + adelantos
        neto = total_ingresos - total_descuentos
        
        resultado['empleados'].append({
            'empleado': empleado,
            'sueldo_base': sueldo_base,
            'sueldo_ajustado': sueldo_ajustado,
            'vacaciones': vacaciones,
            'cts': cts,
            'gratificacion': gratificacion,
            'asignacion_familiar': asignacion_familiar,
            'pension': pension,
            'impuesto_renta': impuesto_renta,
            'descuento_alimentos': descuento_alimentos,
            'descuentos': {
                'prestamos': prestamos,
                'adelantos': adelantos
            },
            'total_ingresos': total_ingresos,
            'total_descuentos': total_descuentos,
            'neto': neto
        })
        
        resultado['totales']['total_ingresos'] += total_ingresos
        resultado['totales']['total_descuentos'] += total_descuentos
        resultado['totales']['total_neto'] += neto
    
    # Calcular para locadores
    for locador in locadores:
        monto_base = locador.monto_mensual
        retencion_4ta = monto_base * 0.08 if not locador.suspendido and monto_base > 1500 else 0
        
        # Descuento por alimentos (solo en pago mensual)
        descuento_alimentos = locador.descuento_alimentos
        
        total_ingresos = monto_base
        total_descuentos = retencion_4ta + descuento_alimentos
        neto = total_ingresos - total_descuentos
        
        resultado['locadores'].append({
            'locador': locador,
            'monto_base': monto_base,
            'retencion_4ta': retencion_4ta,
            'descuento_alimentos': descuento_alimentos,
            'total_ingresos': total_ingresos,
            'total_descuentos': total_descuentos,
            'neto': neto
        })
        
        resultado['totales']['total_ingresos'] += total_ingresos
        resultado['totales']['total_descuentos'] += total_descuentos
        resultado['totales']['total_neto'] += neto
    
    return resultado

@app.route('/cargar_excel/<int:empresa_id>', methods=['GET', 'POST'])
def cargar_excel(empresa_id):
    """Cargar personal desde archivo Excel"""
    empresa = Empresa.query.get_or_404(empresa_id)
    
    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('cargar_excel', empresa_id=empresa_id))
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('cargar_excel', empresa_id=empresa_id))
        
        if archivo and archivo.filename.endswith('.xlsx'):
            try:
                wb = load_workbook(archivo)
                ws = wb.active
                
                empleados_creados = 0
                locadores_creados = 0
                
                # Leer datos de empleados (asumiendo que están en las primeras filas)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:  # Si la primera celda está vacía, terminar
                        break
                    
                    tipo = row[0].lower() if row[0] else ''
                    
                    if tipo == 'empleado':
                        empleado = Empleado(
                            empresa_id=empresa_id,
                            nombres=row[1] or '',
                            apellidos=row[2] or '',
                            dni=row[3] or '',
                            sueldo_base=float(row[4]) if row[4] else 0,
                            fecha_ingreso=row[5] if row[5] else date.today(),
                            fecha_nacimiento=row[6] if row[6] else None,
                            direccion=row[7] or '',
                            telefono=row[8] or '',
                            email=row[9] or '',
                            tipo_pension=row[10] or 'ONP',
                            afp_codigo=row[11] if row[11] else None,
                            cuenta_bancaria=row[12] or '',
                            banco=row[13] or '',
                            tipo_pago=row[14] or 'mensual',
                            descuento_alimentos=float(row[15]) if row[15] else 0
                        )
                        db.session.add(empleado)
                        empleados_creados += 1
                    
                    elif tipo == 'locador':
                        locador = Locador(
                            empresa_id=empresa_id,
                            nombres=row[1] or '',
                            apellidos=row[2] or '',
                            dni=row[3] or '',
                            monto_mensual=float(row[4]) if row[4] else 0,
                            fecha_inicio=row[5] if row[5] else date.today(),
                            suspendido=bool(row[6]) if row[6] else False,
                            cuenta_bancaria=row[7] or '',
                            banco=row[8] or '',
                            descuento_alimentos=float(row[9]) if row[9] else 0
                        )
                        db.session.add(locador)
                        locadores_creados += 1
                
                db.session.commit()
                flash(f'Carga exitosa: {empleados_creados} empleados y {locadores_creados} locadores', 'success')
                return redirect(url_for('personal', empresa_id=empresa_id))
                
            except Exception as e:
                db.session.rollback()
                flash(f'Error al procesar el archivo: {str(e)}', 'error')
        else:
            flash('Formato de archivo no válido. Use archivos .xlsx', 'error')
    
    return render_template('cargar_excel.html', empresa=empresa)

@app.route('/descargar_plantilla/<int:empresa_id>')
def descargar_plantilla(empresa_id):
    """Descargar plantilla Excel para carga masiva"""
    empresa = Empresa.query.get_or_404(empresa_id)
    
    # Crear libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Carga Masiva"
    
    # Encabezados
    headers = [
        'Tipo', 'Nombres', 'Apellidos', 'DNI', 'Sueldo/Monto', 'Fecha Ingreso/Inicio',
        'Fecha Nacimiento', 'Dirección', 'Teléfono', 'Email', 'Tipo Pensión',
        'Código AFP', 'Cuenta Bancaria', 'Banco', 'Tipo Pago', 'Descuento Alimentos'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos de ejemplo para empleados
    ws.cell(row=2, column=1, value='empleado')
    ws.cell(row=2, column=2, value='Juan')
    ws.cell(row=2, column=3, value='Pérez')
    ws.cell(row=2, column=4, value='12345678')
    ws.cell(row=2, column=5, value=1500)
    ws.cell(row=2, column=6, value=date.today())
    ws.cell(row=2, column=7, value=date(1990, 1, 1))
    ws.cell(row=2, column=8, value='Dirección ejemplo')
    ws.cell(row=2, column=9, value='987654321')
    ws.cell(row=2, column=10, value='juan@email.com')
    ws.cell(row=2, column=11, value='ONP')
    ws.cell(row=2, column=12, value='')
    ws.cell(row=2, column=13, value='1234567890123456')
    ws.cell(row=2, column=14, value='BCP')
    ws.cell(row=2, column=15, value='mensual')
    ws.cell(row=2, column=16, value=0)
    
    # Datos de ejemplo para locadores
    ws.cell(row=3, column=1, value='locador')
    ws.cell(row=3, column=2, value='María')
    ws.cell(row=3, column=3, value='García')
    ws.cell(row=3, column=4, value='87654321')
    ws.cell(row=3, column=5, value=3000)
    ws.cell(row=3, column=6, value=date.today())
    ws.cell(row=3, column=7, value='')
    ws.cell(row=3, column=8, value='')
    ws.cell(row=3, column=9, value='')
    ws.cell(row=3, column=10, value='')
    ws.cell(row=3, column=11, value='')
    ws.cell(row=3, column=12, value='')
    ws.cell(row=3, column=13, value='9876543210987654')
    ws.cell(row=3, column=14, value='BBVA')
    ws.cell(row=3, column=15, value='')
    ws.cell(row=3, column=16, value=0)
    
    # Crear respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'plantilla_carga_{empresa.nombre}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000)
