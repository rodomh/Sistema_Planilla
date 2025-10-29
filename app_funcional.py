from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
import os
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['SECRET_KEY'] = 'sispla_peru_2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///sispla.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Crear directorio de uploads si no existe
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)

# Modelos que coinciden exactamente con la base de datos existente
class Empresa(db.Model):
    __tablename__ = 'empresas'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(200), nullable=False)
    ruc = db.Column(db.String(11), unique=True, nullable=False)
    regimen_laboral = db.Column(db.String(50), nullable=False)
    direccion = db.Column(db.String(500))
    telefono = db.Column(db.String(20))
    email = db.Column(db.String(100))
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    activa = db.Column(db.Boolean, default=True)

class Empleado(db.Model):
    __tablename__ = 'empleados'
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresas.id'), nullable=False)
    nombres = db.Column(db.String(100), nullable=False)
    apellidos = db.Column(db.String(100), nullable=False)
    dni = db.Column(db.String(8), nullable=False)
    sueldo_base = db.Column(db.Float, nullable=False)
    fecha_ingreso = db.Column(db.Date, nullable=False)
    fecha_nacimiento = db.Column(db.Date)
    direccion = db.Column(db.String(500))
    telefono = db.Column(db.String(20))
    email = db.Column(db.String(100))
    tipo_pension = db.Column(db.String(20), default='ONP')
    afp_codigo = db.Column(db.String(10))
    cuenta_bancaria = db.Column(db.String(20))
    banco = db.Column(db.String(50))
    tipo_pago = db.Column(db.String(20), default='mensual')
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

class Locador(db.Model):
    __tablename__ = 'locadores'
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresas.id'), nullable=False)
    nombres = db.Column(db.String(100), nullable=False)
    apellidos = db.Column(db.String(100), nullable=False)
    dni = db.Column(db.String(8), nullable=False)
    monto_mensual = db.Column(db.Float, nullable=False)
    fecha_inicio = db.Column(db.Date, nullable=False)
    fecha_fin = db.Column(db.Date)
    direccion = db.Column(db.String(500))
    telefono = db.Column(db.String(20))
    email = db.Column(db.String(100))
    cuenta_bancaria = db.Column(db.String(20))
    banco = db.Column(db.String(50))
    suspendido = db.Column(db.Boolean, default=False)
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

# Rutas principales
@app.route('/')
def index():
    """Página principal del sistema"""
    empresas = Empresa.query.all()
    
    # Calcular conteos de empleados y locadores activos por empresa
    empresas_con_conteos = []
    for empresa in empresas:
        empleados_activos = Empleado.query.filter_by(empresa_id=empresa.id, activo=True).count()
        locadores_activos = Locador.query.filter_by(empresa_id=empresa.id, activo=True).count()
        
        empresas_con_conteos.append({
            'empresa': empresa,
            'empleados_activos': empleados_activos,
            'locadores_activos': locadores_activos
        })
    
    return render_template('index.html', empresas_con_conteos=empresas_con_conteos)

@app.route('/empresas')
def empresas():
    """Gestión de empresas"""
    empresas = Empresa.query.all()
    return render_template('empresas.html', empresas=empresas)

@app.route('/empresa/nueva', methods=['GET', 'POST'])
def nueva_empresa():
    """Crear nueva empresa"""
    if request.method == 'POST':
        nombre = request.form['nombre']
        regimen = request.form['regimen_laboral']
        ruc = request.form['ruc']
        
        empresa = Empresa(
            nombre=nombre,
            regimen_laboral=regimen,
            ruc=ruc
        )
        db.session.add(empresa)
        db.session.commit()
        flash('Empresa creada exitosamente', 'success')
        return redirect(url_for('empresas'))
    
    return render_template('nueva_empresa.html')

@app.route('/personal/<int:empresa_id>')
def personal(empresa_id):
    """Gestión de personal por empresa"""
    empresa = Empresa.query.get_or_404(empresa_id)
    empleados = Empleado.query.filter_by(empresa_id=empresa_id).all()
    locadores = Locador.query.filter_by(empresa_id=empresa_id).all()
    return render_template('personal.html', empresa=empresa, empleados=empleados, locadores=locadores)

@app.route('/empleado/nuevo/<int:empresa_id>', methods=['GET', 'POST'])
def nuevo_empleado(empresa_id):
    """Crear nuevo empleado"""
    empresa = Empresa.query.get_or_404(empresa_id)
    
    if request.method == 'POST':
        empleado = Empleado(
            empresa_id=empresa_id,
            nombres=request.form['nombres'],
            apellidos=request.form['apellidos'],
            dni=request.form['dni'],
            sueldo_base=float(request.form['sueldo_base']),
            fecha_ingreso=datetime.strptime(request.form['fecha_ingreso'], '%Y-%m-%d').date(),
            tipo_pension=request.form['tipo_pension'],
            afp_codigo=request.form.get('afp_codigo', '')
        )
        db.session.add(empleado)
        db.session.commit()
        flash('Empleado creado exitosamente', 'success')
        return redirect(url_for('personal', empresa_id=empresa_id))
    
    return render_template('nuevo_empleado.html', empresa=empresa)

@app.route('/locador/nuevo/<int:empresa_id>', methods=['GET', 'POST'])
def nuevo_locador(empresa_id):
    """Crear nuevo locador de servicios"""
    empresa = Empresa.query.get_or_404(empresa_id)
    
    if request.method == 'POST':
        locador = Locador(
            empresa_id=empresa_id,
            nombres=request.form['nombres'],
            apellidos=request.form['apellidos'],
            dni=request.form['dni'],
            monto_mensual=float(request.form['monto_mensual']),
            fecha_inicio=datetime.strptime(request.form['fecha_inicio'], '%Y-%m-%d').date(),
            suspendido=request.form.get('suspendido') == 'on'
        )
        db.session.add(locador)
        db.session.commit()
        flash('Locador creado exitosamente', 'success')
        return redirect(url_for('personal', empresa_id=empresa_id))
    
    return render_template('nuevo_locador.html', empresa=empresa)

@app.route('/empleado/editar/<int:empleado_id>', methods=['GET', 'POST'])
def editar_empleado(empleado_id):
    """Editar empleado existente"""
    empleado = Empleado.query.get_or_404(empleado_id)
    empresa = Empresa.query.get_or_404(empleado.empresa_id)
    
    if request.method == 'POST':
        empleado.nombres = request.form['nombres']
        empleado.apellidos = request.form['apellidos']
        empleado.dni = request.form['dni']
        empleado.sueldo_base = float(request.form['sueldo_base'])
        empleado.fecha_ingreso = datetime.strptime(request.form['fecha_ingreso'], '%Y-%m-%d').date()
        empleado.tipo_pension = request.form['tipo_pension']
        empleado.afp_codigo = request.form.get('afp_codigo', '')
        empleado.cuenta_bancaria = request.form.get('cuenta_bancaria', '')
        empleado.banco = request.form.get('banco', '')
        empleado.tipo_pago = request.form.get('tipo_pago', 'mensual')
        empleado.activo = request.form.get('activo') == 'on'
        
        db.session.commit()
        flash('Empleado actualizado exitosamente', 'success')
        return redirect(url_for('personal', empresa_id=empresa.id))
    
    return render_template('editar_empleado.html', empleado=empleado, empresa=empresa)

@app.route('/locador/editar/<int:locador_id>', methods=['GET', 'POST'])
def editar_locador(locador_id):
    """Editar locador existente"""
    locador = Locador.query.get_or_404(locador_id)
    empresa = Empresa.query.get_or_404(locador.empresa_id)
    
    if request.method == 'POST':
        locador.nombres = request.form['nombres']
        locador.apellidos = request.form['apellidos']
        locador.dni = request.form['dni']
        locador.monto_mensual = float(request.form['monto_mensual'])
        locador.fecha_inicio = datetime.strptime(request.form['fecha_inicio'], '%Y-%m-%d').date()
        locador.suspendido = request.form.get('suspendido') == 'on'
        locador.activo = request.form.get('activo') == 'on'
        locador.cuenta_bancaria = request.form.get('cuenta_bancaria', '')
        locador.banco = request.form.get('banco', '')
        
        db.session.commit()
        flash('Locador actualizado exitosamente', 'success')
        return redirect(url_for('personal', empresa_id=empresa.id))
    
    return render_template('editar_locador.html', locador=locador, empresa=empresa)

@app.route('/empleado/eliminar/<int:empleado_id>', methods=['POST'])
def eliminar_empleado(empleado_id):
    """Eliminar empleado (marcar como inactivo)"""
    empleado = Empleado.query.get_or_404(empleado_id)
    empresa_id = empleado.empresa_id
    empleado.activo = False
    db.session.commit()
    flash('Empleado eliminado exitosamente', 'success')
    return redirect(url_for('personal', empresa_id=empresa_id))

@app.route('/locador/eliminar/<int:locador_id>', methods=['POST'])
def eliminar_locador(locador_id):
    """Eliminar locador (marcar como inactivo)"""
    locador = Locador.query.get_or_404(locador_id)
    empresa_id = locador.empresa_id
    locador.activo = False
    db.session.commit()
    flash('Locador eliminado exitosamente', 'success')
    return redirect(url_for('personal', empresa_id=empresa_id))

@app.route('/cargar_excel/<int:empresa_id>', methods=['GET', 'POST'])
def cargar_excel(empresa_id):
    """Cargar personal desde archivo Excel"""
    empresa = Empresa.query.get_or_404(empresa_id)
    
    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(request.url)
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(request.url)
        
        if archivo and archivo.filename.endswith('.xlsx'):
            try:
                # Guardar archivo temporalmente
                filename = secure_filename(archivo.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                archivo.save(filepath)
                
                # Procesar archivo Excel
                wb = load_workbook(filepath)
                
                # Procesar empleados si existe la hoja
                if 'Empleados' in wb.sheetnames:
                    ws_empleados = wb['Empleados']
                    empleados_cargados = 0
                    
                    for row in ws_empleados.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1] and row[2]:  # Nombres, Apellidos, DNI
                            empleado = Empleado(
                                empresa_id=empresa_id,
                                nombres=str(row[0]),
                                apellidos=str(row[1]),
                                dni=str(row[2]),
                                sueldo_base=float(row[3]) if row[3] else 0,
                                fecha_ingreso=row[4] if row[4] else date.today(),
                                tipo_pension=str(row[5]) if row[5] else 'ONP',
                                afp_codigo=str(row[6]) if row[6] else '',
                                cuenta_bancaria=str(row[7]) if row[7] else '',
                                banco=str(row[8]) if row[8] else '',
                                tipo_pago=str(row[9]) if row[9] else 'mensual',
                                activo=True
                            )
                            db.session.add(empleado)
                            empleados_cargados += 1
                    
                    db.session.commit()
                    flash(f'Se cargaron {empleados_cargados} empleados exitosamente', 'success')
                
                # Procesar locadores si existe la hoja
                if 'Locadores' in wb.sheetnames:
                    ws_locadores = wb['Locadores']
                    locadores_cargados = 0
                    
                    for row in ws_locadores.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1] and row[2]:  # Nombres, Apellidos, DNI
                            locador = Locador(
                                empresa_id=empresa_id,
                                nombres=str(row[0]),
                                apellidos=str(row[1]),
                                dni=str(row[2]),
                                monto_mensual=float(row[3]) if row[3] else 0,
                                fecha_inicio=row[4] if row[4] else date.today(),
                                suspendido=bool(row[5]) if row[5] is not None else False,
                                cuenta_bancaria=str(row[6]) if row[6] else '',
                                banco=str(row[7]) if row[7] else '',
                                activo=True
                            )
                            db.session.add(locador)
                            locadores_cargados += 1
                    
                    db.session.commit()
                    flash(f'Se cargaron {locadores_cargados} locadores exitosamente', 'success')
                
                # Eliminar archivo temporal
                os.remove(filepath)
                
                return redirect(url_for('personal', empresa_id=empresa_id))
                
            except Exception as e:
                flash(f'Error al procesar el archivo: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('El archivo debe ser un Excel (.xlsx)', 'error')
            return redirect(request.url)
    
    return render_template('cargar_excel.html', empresa=empresa)

@app.route('/descargar_plantilla/<int:empresa_id>')
def descargar_plantilla(empresa_id):
    """Descargar plantilla Excel para carga de personal"""
    empresa = Empresa.query.get_or_404(empresa_id)
    
    # Crear libro de trabajo
    wb = Workbook()
    
    # Hoja de empleados
    ws_empleados = wb.active
    ws_empleados.title = "Empleados"
    
    # Encabezados para empleados
    headers_empleados = [
        'Nombres', 'Apellidos', 'DNI', 'Sueldo Base', 'Fecha Ingreso',
        'Tipo Pensión', 'AFP Código', 'Cuenta Bancaria', 'Banco', 'Tipo Pago'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers_empleados, 1):
        cell = ws_empleados.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Agregar fila de ejemplo
    ejemplo_empleado = [
        'Juan', 'Pérez García', '12345678', 1500.00, '2024-01-15',
        'ONP', '', '1234567890123456', 'BCP', 'mensual'
    ]
    
    for col, valor in enumerate(ejemplo_empleado, 1):
        ws_empleados.cell(row=2, column=col, value=valor)
    
    # Hoja de locadores
    ws_locadores = wb.create_sheet("Locadores")
    
    # Encabezados para locadores
    headers_locadores = [
        'Nombres', 'Apellidos', 'DNI', 'Monto Mensual', 'Fecha Inicio',
        'Suspendido', 'Cuenta Bancaria', 'Banco'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers_locadores, 1):
        cell = ws_locadores.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Agregar fila de ejemplo
    ejemplo_locador = [
        'María', 'Rodríguez López', '87654321', 2000.00, '2024-02-01',
        False, '9876543210987654', 'BBVA'
    ]
    
    for col, valor in enumerate(ejemplo_locador, 1):
        ws_locadores.cell(row=2, column=col, value=valor)
    
    # Ajustar ancho de columnas
    for ws in [ws_empleados, ws_locadores]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=plantilla_carga_{empresa.nombre}.xlsx'
    
    return response

@app.route('/planilla/<int:empresa_id>')
def planilla(empresa_id):
    """Vista de planilla por empresa"""
    empresa = Empresa.query.get_or_404(empresa_id)
    empleados = Empleado.query.filter_by(empresa_id=empresa_id).all()
    locadores = Locador.query.filter_by(empresa_id=empresa_id).all()
    
    # Obtener mes actual
    mes_actual = datetime.now().month
    año_actual = datetime.now().year
    
    return render_template('planilla.html', 
                         empresa=empresa, 
                         empleados=empleados, 
                         locadores=locadores,
                         mes=mes_actual,
                         año=año_actual)

def calcular_planilla_simple(empresa_id, mes, año):
    """Función simplificada para calcular planilla"""
    empresa = Empresa.query.get_or_404(empresa_id)
    empleados = Empleado.query.filter_by(empresa_id=empresa_id, activo=True).all()
    locadores = Locador.query.filter_by(empresa_id=empresa_id, activo=True).all()
    
    resultados = {
        'empresa': empresa.nombre,
        'regimen_laboral': empresa.regimen_laboral,
        'mes': mes,
        'año': año,
        'empleados': [],
        'locadores': [],
        'totales': {
            'total_empleados': 0,
            'total_locadores': 0,
            'total_ingresos': 0,
            'total_descuentos': 0,
            'total_neto': 0
        }
    }
    
    # Calcular planillas de empleados
    for empleado in empleados:
        # Cálculo básico
        sueldo_base = float(empleado.sueldo_base)
        sueldo_ajustado = sueldo_base
        
        # Beneficios según régimen
        if empresa.regimen_laboral in ['microempresa', 'pequeña_empresa']:
            vacaciones = sueldo_ajustado * 15 / 360
        else:
            vacaciones = sueldo_ajustado * 30 / 360
            
        if empresa.regimen_laboral == 'microempresa':
            cts = 0
        elif empresa.regimen_laboral == 'pequeña_empresa':
            cts = sueldo_ajustado * 15 / 12
        else:
            cts = sueldo_ajustado
            
        if empresa.regimen_laboral == 'microempresa':
            gratificacion = 0
        elif empresa.regimen_laboral == 'pequeña_empresa' and mes in [7, 12]:
            gratificacion = sueldo_ajustado / 2
        elif empresa.regimen_laboral == 'general' and mes in [7, 12]:
            gratificacion = sueldo_ajustado * 1.09
        else:
            gratificacion = 0
            
        if empresa.regimen_laboral in ['pequeña_empresa', 'general'] and sueldo_ajustado <= 1025:
            asignacion_familiar = 102.50
        else:
            asignacion_familiar = 0
        
        # Descuentos
        if empleado.tipo_pension == 'ONP':
            pension = sueldo_ajustado * 0.13
        else:
            pension = sueldo_ajustado * 0.12
            
        if sueldo_ajustado > 1025:
            impuesto_renta = max(0, (sueldo_ajustado - 1025) * 0.08)
        else:
            impuesto_renta = 0
        
        # Totales
        total_ingresos = sueldo_ajustado + vacaciones + cts + gratificacion + asignacion_familiar
        total_descuentos = pension + impuesto_renta
        neto_pagar = total_ingresos - total_descuentos
        
        resultados['empleados'].append({
            'id': empleado.id,
            'nombres': empleado.nombres,
            'apellidos': empleado.apellidos,
            'dni': empleado.dni,
            'sueldo_base': sueldo_base,
            'sueldo_ajustado': round(sueldo_ajustado, 2),
            'dias_trabajados': 30,
            'dias_faltados': 0,
            'horas_perdidas': 0,
            'beneficios': {
                'vacaciones': round(vacaciones, 2),
                'cts': round(cts, 2),
                'gratificacion': round(gratificacion, 2),
                'asignacion_familiar': round(asignacion_familiar, 2)
            },
            'descuentos': {
                'pension': round(pension, 2),
                'impuesto_renta': round(impuesto_renta, 2),
                'prestamos': 0,
                'adelantos': 0
            },
            'total_ingresos': round(total_ingresos, 2),
            'total_descuentos': round(total_descuentos, 2),
            'neto_pagar': round(neto_pagar, 2)
        })
        
        resultados['totales']['total_ingresos'] += total_ingresos
        resultados['totales']['total_descuentos'] += total_descuentos
        resultados['totales']['total_neto'] += neto_pagar
    
    # Calcular planillas de locadores
    for locador in locadores:
        monto_bruto = float(locador.monto_mensual)
        if not locador.suspendido and monto_bruto > 1500:
            retencion_4ta_cat = monto_bruto * 0.08
        else:
            retencion_4ta_cat = 0
        neto_pagar = monto_bruto - retencion_4ta_cat
        
        resultados['locadores'].append({
            'id': locador.id,
            'nombres': locador.nombres,
            'apellidos': locador.apellidos,
            'dni': locador.dni,
            'monto_mensual': monto_bruto,
            'suspendido': locador.suspendido,
            'monto_bruto': monto_bruto,
            'retencion_4ta_cat': round(retencion_4ta_cat, 2),
            'neto_pagar': round(neto_pagar, 2)
        })
        
        resultados['totales']['total_neto'] += neto_pagar
    
    # Actualizar contadores
    resultados['totales']['total_empleados'] = len(empleados)
    resultados['totales']['total_locadores'] = len(locadores)
    resultados['totales']['total_ingresos'] = round(resultados['totales']['total_ingresos'], 2)
    resultados['totales']['total_descuentos'] = round(resultados['totales']['total_descuentos'], 2)
    resultados['totales']['total_neto'] = round(resultados['totales']['total_neto'], 2)
    
    return resultados

@app.route('/calcular_planilla/<int:empresa_id>', methods=['POST'])
def calcular_planilla(empresa_id):
    """Calcular planilla para una empresa específica"""
    empresa = Empresa.query.get_or_404(empresa_id)
    mes = int(request.form['mes'])
    año = int(request.form['año'])
    
    try:
        resultado = calcular_planilla_simple(empresa_id, mes, año)
        return jsonify({
            'success': True,
            'resultado': resultado,
            'empresa': empresa.nombre,
            'mes': mes,
            'año': año
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/exportar_excel/<int:empresa_id>', methods=['POST'])
def exportar_excel(empresa_id):
    """Exportar planilla a Excel"""
    empresa = Empresa.query.get_or_404(empresa_id)
    mes = int(request.form['mes'])
    año = int(request.form['año'])
    
    try:
        resultado = calcular_planilla_simple(empresa_id, mes, año)
        
        # Crear libro de trabajo
        wb = Workbook()
        
        # Hoja de empleados
        ws_empleados = wb.active
        ws_empleados.title = "Empleados"
        
        # Encabezados para empleados
        headers_empleados = [
            'Nombres', 'Apellidos', 'DNI', 'Sueldo Base', 'Sueldo Ajustado',
            'Días Trabajados', 'Días Faltados', 'Horas Perdidas',
            'Vacaciones', 'CTS', 'Gratificación', 'Asignación Familiar',
            'Pensión', 'Impuesto Renta', 'Préstamos', 'Adelantos', 'Faltas',
            'Total Ingresos', 'Total Descuentos', 'Neto a Pagar'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers_empleados, 1):
            cell = ws_empleados.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Escribir datos de empleados
        for row, emp in enumerate(resultado['empleados'], 2):
            ws_empleados.cell(row=row, column=1, value=emp['nombres'])
            ws_empleados.cell(row=row, column=2, value=emp['apellidos'])
            ws_empleados.cell(row=row, column=3, value=emp['dni'])
            ws_empleados.cell(row=row, column=4, value=emp['sueldo_base'])
            ws_empleados.cell(row=row, column=5, value=emp['sueldo_ajustado'])
            ws_empleados.cell(row=row, column=6, value=emp['dias_trabajados'])
            ws_empleados.cell(row=row, column=7, value=emp['dias_faltados'])
            ws_empleados.cell(row=row, column=8, value=emp['horas_perdidas'])
            ws_empleados.cell(row=row, column=9, value=emp['beneficios']['vacaciones'])
            ws_empleados.cell(row=row, column=10, value=emp['beneficios']['cts'])
            ws_empleados.cell(row=row, column=11, value=emp['beneficios']['gratificacion'])
            ws_empleados.cell(row=row, column=12, value=emp['beneficios']['asignacion_familiar'])
            ws_empleados.cell(row=row, column=13, value=emp['descuentos']['pension'])
            ws_empleados.cell(row=row, column=14, value=emp['descuentos']['impuesto_renta'])
            ws_empleados.cell(row=row, column=15, value=emp['descuentos']['prestamos'])
            ws_empleados.cell(row=row, column=16, value=emp['descuentos']['adelantos'])
            ws_empleados.cell(row=row, column=17, value=emp.get('faltas', 0))
            ws_empleados.cell(row=row, column=18, value=emp['total_ingresos'])
            ws_empleados.cell(row=row, column=19, value=emp['total_descuentos'])
            ws_empleados.cell(row=row, column=20, value=emp['neto_pagar'])
        
        # Ajustar ancho de columnas
        for column in ws_empleados.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_empleados.column_dimensions[column_letter].width = adjusted_width
        
        # Hoja de locadores
        ws_locadores = wb.create_sheet("Locadores")
        
        # Encabezados para locadores
        headers_locadores = [
            'Nombres', 'Apellidos', 'DNI', 'Monto Mensual', 'Suspendido',
            'Retención 4ta Cat', 'Neto a Pagar'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers_locadores, 1):
            cell = ws_locadores.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Escribir datos de locadores
        for row, loc in enumerate(resultado['locadores'], 2):
            ws_locadores.cell(row=row, column=1, value=loc['nombres'])
            ws_locadores.cell(row=row, column=2, value=loc['apellidos'])
            ws_locadores.cell(row=row, column=3, value=loc['dni'])
            ws_locadores.cell(row=row, column=4, value=loc['monto_mensual'])
            ws_locadores.cell(row=row, column=5, value='Sí' if loc['suspendido'] else 'No')
            ws_locadores.cell(row=row, column=6, value=loc['retencion_4ta_cat'])
            ws_locadores.cell(row=row, column=7, value=loc['neto_pagar'])
        
        # Ajustar ancho de columnas
        for column in ws_locadores.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_locadores.column_dimensions[column_letter].width = adjusted_width
        
        # Hoja de resumen
        ws_resumen = wb.create_sheet("Resumen")
        
        # Escribir resumen
        ws_resumen.cell(row=1, column=1, value="Concepto").font = Font(bold=True)
        ws_resumen.cell(row=1, column=2, value="Valor").font = Font(bold=True)
        
        resumen_data = [
            ("Empresa", resultado['empresa']),
            ("Régimen Laboral", resultado['regimen_laboral']),
            ("Período", f"{mes:02d}/{año}"),
            ("Total Empleados", resultado['totales']['total_empleados']),
            ("Total Locadores", resultado['totales']['total_locadores']),
            ("Total Ingresos", resultado['totales']['total_ingresos']),
            ("Total Descuentos", resultado['totales']['total_descuentos']),
            ("Total Neto", resultado['totales']['total_neto'])
        ]
        
        for row, (concepto, valor) in enumerate(resumen_data, 2):
            ws_resumen.cell(row=row, column=1, value=concepto)
            ws_resumen.cell(row=row, column=2, value=valor)
        
        # Ajustar ancho de columnas
        for column in ws_resumen.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_resumen.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=planilla_{empresa.nombre}_{mes:02d}_{año}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Error al exportar: {str(e)}', 'error')
        return redirect(url_for('planilla', empresa_id=empresa_id))

if __name__ == '__main__':
    # No crear tablas, usar las existentes
    app.run(debug=True, host='0.0.0.0', port=5000)
