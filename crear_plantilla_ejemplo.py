"""
Script para crear una plantilla de ejemplo con datos de muestra
para la carga masiva de personal y locadores
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import date, timedelta
import random

def crear_plantilla_ejemplo():
    """Crear una plantilla de ejemplo con datos de muestra"""
    print("Creando plantilla de ejemplo con datos de muestra...")
    
    # Crear libro de trabajo
    wb = Workbook()
    
    # Hoja de empleados
    ws_empleados = wb.active
    ws_empleados.title = "Empleados"
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados para empleados
    headers_empleados = [
        'Nombres', 'Apellidos', 'DNI', 'Sueldo Base', 'Fecha Ingreso',
        'Tipo Pensión', 'AFP Código', 'Cuenta Bancaria', 'Banco', 'Tipo Pago'
    ]
    
    # Escribir encabezados con estilo
    for col, header in enumerate(headers_empleados, 1):
        cell = ws_empleados.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Datos de ejemplo para empleados
    nombres_ejemplo = ['Juan', 'María', 'Carlos', 'Ana', 'Luis', 'Elena', 'Roberto', 'Patricia']
    apellidos_ejemplo = ['Pérez', 'López', 'García', 'Rodríguez', 'Martínez', 'González', 'Herrera', 'Díaz']
    bancos_ejemplo = ['BCP', 'BBVA', 'SCOTIABANK', 'INTERBANK', 'BANBIF', 'BANCO PICHINCHA']
    afps_ejemplo = ['PRIMA', 'HABITAT', 'PROFUTURO', 'INTEGRA']
    
    # Generar datos de ejemplo
    for row in range(2, 12):  # 10 empleados de ejemplo
        # Generar DNI único
        dni = f"{random.randint(10000000, 99999999)}"
        
        # Generar fecha de ingreso (últimos 12 meses)
        fecha_ingreso = date.today() - timedelta(days=random.randint(30, 365))
        
        # Generar sueldo base
        sueldo_base = round(random.uniform(1200, 3500), 2)
        
        # Seleccionar tipo de pensión
        tipo_pension = random.choice(['ONP', 'AFP'])
        afp_codigo = random.choice(afps_ejemplo) if tipo_pension == 'AFP' else ''
        
        # Generar cuenta bancaria
        cuenta_bancaria = f"{random.randint(1000000000000000, 9999999999999999)}"
        
        # Seleccionar banco
        banco = random.choice(bancos_ejemplo)
        
        # Seleccionar tipo de pago
        tipo_pago = random.choice(['mensual', 'quincenal'])
        
        # Datos del empleado
        empleado_data = [
            random.choice(nombres_ejemplo),
            random.choice(apellidos_ejemplo),
            dni,
            sueldo_base,
            fecha_ingreso.strftime('%Y-%m-%d'),
            tipo_pension,
            afp_codigo,
            cuenta_bancaria,
            banco,
            tipo_pago
        ]
        
        # Escribir datos
        for col, valor in enumerate(empleado_data, 1):
            cell = ws_empleados.cell(row=row, column=col, value=valor)
            cell.border = border
            if col in [4, 7]:  # Sueldo Base y Cuenta Bancaria
                cell.alignment = Alignment(horizontal="right")
    
    # Ajustar ancho de columnas
    for column in ws_empleados.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_empleados.column_dimensions[column_letter].width = adjusted_width
    
    # Hoja de locadores
    ws_locadores = wb.create_sheet("Locadores")
    
    # Encabezados para locadores
    headers_locadores = [
        'Nombres', 'Apellidos', 'DNI', 'Monto Mensual', 'Fecha Inicio',
        'Suspendido', 'Cuenta Bancaria', 'Banco'
    ]
    
    # Escribir encabezados con estilo
    for col, header in enumerate(headers_locadores, 1):
        cell = ws_locadores.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Generar datos de ejemplo para locadores
    for row in range(2, 8):  # 6 locadores de ejemplo
        # Generar DNI único
        dni = f"{random.randint(10000000, 99999999)}"
        
        # Generar fecha de inicio
        fecha_inicio = date.today() - timedelta(days=random.randint(30, 365))
        
        # Generar monto mensual
        monto_mensual = round(random.uniform(2000, 5000), 2)
        
        # Seleccionar si está suspendido
        suspendido = random.choice([True, False])
        
        # Generar cuenta bancaria
        cuenta_bancaria = f"{random.randint(1000000000000000, 9999999999999999)}"
        
        # Seleccionar banco
        banco = random.choice(bancos_ejemplo)
        
        # Datos del locador
        locador_data = [
            random.choice(nombres_ejemplo),
            random.choice(apellidos_ejemplo),
            dni,
            monto_mensual,
            fecha_inicio.strftime('%Y-%m-%d'),
            suspendido,
            cuenta_bancaria,
            banco
        ]
        
        # Escribir datos
        for col, valor in enumerate(locador_data, 1):
            cell = ws_locadores.cell(row=row, column=col, value=valor)
            cell.border = border
            if col in [4, 7]:  # Monto Mensual y Cuenta Bancaria
                cell.alignment = Alignment(horizontal="right")
    
    # Ajustar ancho de columnas para locadores
    for column in ws_locadores.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_locadores.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar archivo
    filename = 'PLANTILLA_CARGA_MASIVA_EJEMPLO.xlsx'
    wb.save(filename)
    print(f"✅ Plantilla de ejemplo creada: {filename}")
    print(f"   - 10 empleados de ejemplo")
    print(f"   - 6 locadores de ejemplo")
    print(f"   - Datos realistas generados automáticamente")
    print(f"   - Formato profesional con estilos")
    
    return filename

if __name__ == '__main__':
    print("="*60)
    print("CREADOR DE PLANTILLA DE EJEMPLO PARA CARGA MASIVA")
    print("="*60)
    
    try:
        archivo = crear_plantilla_ejemplo()
        print("\n" + "="*60)
        print("PLANTILLA CREADA EXITOSAMENTE")
        print("="*60)
        print("Puedes usar este archivo como:")
        print("1. Plantilla para crear tus propios archivos de carga")
        print("2. Ejemplo de la estructura correcta")
        print("3. Archivo de prueba para la carga masiva")
        print("\nEl archivo está listo para usar en el sistema.")
        
    except Exception as e:
        print(f"❌ Error al crear la plantilla: {e}")
        import traceback
        traceback.print_exc()
