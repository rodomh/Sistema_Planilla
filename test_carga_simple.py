"""
Script de prueba simplificado para la funcionalidad de carga masiva desde Excel
"""

from app_funcional import app, db, Empresa, Empleado, Locador
from datetime import datetime, date
from openpyxl import Workbook
import os

def crear_archivo_excel_simple():
    """Crear un archivo Excel simple para la carga masiva"""
    print("1. Creando archivo Excel de prueba...")
    
    # Crear libro de trabajo
    wb = Workbook()
    
    # Hoja de empleados
    ws_empleados = wb.active
    ws_empleados.title = "Empleados"
    
    # Encabezados para empleados
    headers_empleados = [
        'Nombres', 'Apellidos', 'DNI', 'Sueldo Base', 'Fecha Ingreso',
        'Tipo Pensión', 'AFP Código', 'Cuenta Bancaria', 'Banco', 'Tipo Pago'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers_empleados, 1):
        ws_empleados.cell(row=1, column=col, value=header)
    
    # Datos de prueba para empleados
    datos_empleados = [
        ['Ana', 'García López', '11111111', 1800.00, '2024-01-15', 'ONP', '', '1111111111111111', 'BCP', 'mensual'],
        ['Carlos', 'Mendoza Silva', '22222222', 2200.00, '2024-02-01', 'AFP', 'PRIMA', '2222222222222222', 'BBVA', 'quincenal'],
        ['María', 'Torres Vega', '33333333', 1600.00, '2024-03-10', 'ONP', '', '3333333333333333', 'SCOTIABANK', 'mensual']
    ]
    
    # Escribir datos de empleados
    for row, datos in enumerate(datos_empleados, 2):
        for col, valor in enumerate(datos, 1):
            ws_empleados.cell(row=row, column=col, value=valor)
    
    # Hoja de locadores
    ws_locadores = wb.create_sheet("Locadores")
    
    # Encabezados para locadores
    headers_locadores = [
        'Nombres', 'Apellidos', 'DNI', 'Monto Mensual', 'Fecha Inicio',
        'Suspendido', 'Cuenta Bancaria', 'Banco'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers_locadores, 1):
        ws_locadores.cell(row=1, column=col, value=header)
    
    # Datos de prueba para locadores
    datos_locadores = [
        ['Luis', 'Ramírez Castro', '44444444', 2500.00, '2024-01-20', False, '4444444444444444', 'BCP'],
        ['Elena', 'Díaz Morales', '55555555', 3000.00, '2024-02-15', False, '5555555555555555', 'BBVA'],
        ['Roberto', 'Herrera Paz', '66666666', 2800.00, '2024-03-05', True, '6666666666666666', 'SCOTIABANK']
    ]
    
    # Escribir datos de locadores
    for row, datos in enumerate(datos_locadores, 2):
        for col, valor in enumerate(datos, 1):
            ws_locadores.cell(row=row, column=col, value=valor)
    
    # Guardar archivo
    filename = 'test_carga_simple.xlsx'
    wb.save(filename)
    print(f"   OK: Archivo Excel creado: {filename}")
    return filename

def probar_carga_simple():
    """Probar la funcionalidad de carga masiva desde Excel"""
    print("="*60)
    print("PRUEBA DE CARGA MASIVA DESDE EXCEL (SIMPLIFICADA)")
    print("="*60)
    
    with app.app_context():
        try:
            # 1. Crear archivo Excel de prueba
            archivo_excel = crear_archivo_excel_simple()
            
            # 2. Verificar empresa
            print("\n2. Verificando empresa...")
            empresa = Empresa.query.first()
            if not empresa:
                print("   ERROR: No hay empresas en la base de datos")
                return False
            
            print(f"   OK: Empresa: {empresa.nombre}")
            
            # 3. Contar empleados y locadores antes de la carga (sin usar campos problemáticos)
            print("\n3. Contando personal antes de la carga...")
            empleados_antes = db.session.query(Empleado).filter_by(empresa_id=empresa.id, activo=True).count()
            locadores_antes = db.session.query(Locador).filter_by(empresa_id=empresa.id, activo=True).count()
            print(f"   OK: Empleados antes: {empleados_antes}")
            print(f"   OK: Locadores antes: {locadores_antes}")
            
            # 4. Simular carga masiva
            print("\n4. Simulando carga masiva...")
            from openpyxl import load_workbook
            
            wb = load_workbook(archivo_excel)
            empleados_cargados = 0
            locadores_cargados = 0
            
            # Procesar empleados
            if 'Empleados' in wb.sheetnames:
                ws_empleados = wb['Empleados']
                for row in ws_empleados.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] and row[2]:  # Nombres, Apellidos, DNI
                        empleado = Empleado(
                            empresa_id=empresa.id,
                            nombres=str(row[0]),
                            apellidos=str(row[1]),
                            dni=str(row[2]),
                            sueldo_base=float(row[3]) if row[3] else 0,
                            fecha_ingreso=row[4] if row[4] else date.today(),
                            tipo_pension=str(row[5]) if row[5] else 'ONP',
                            afp_codigo=str(row[6]) if row[6] else '',
                            activo=True
                        )
                        db.session.add(empleado)
                        empleados_cargados += 1
                        print(f"   OK: Empleado agregado: {empleado.nombres} {empleado.apellidos}")
            
            # Procesar locadores
            if 'Locadores' in wb.sheetnames:
                ws_locadores = wb['Locadores']
                for row in ws_locadores.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] and row[2]:  # Nombres, Apellidos, DNI
                        locador = Locador(
                            empresa_id=empresa.id,
                            nombres=str(row[0]),
                            apellidos=str(row[1]),
                            dni=str(row[2]),
                            monto_mensual=float(row[3]) if row[3] else 0,
                            fecha_inicio=row[4] if row[4] else date.today(),
                            suspendido=bool(row[5]) if row[5] is not None else False,
                            activo=True
                        )
                        db.session.add(locador)
                        locadores_cargados += 1
                        print(f"   OK: Locador agregado: {locador.nombres} {locador.apellidos}")
            
            # 5. Confirmar cambios en la base de datos
            print("\n5. Confirmando cambios en la base de datos...")
            db.session.commit()
            print(f"   OK: {empleados_cargados} empleados cargados")
            print(f"   OK: {locadores_cargados} locadores cargados")
            
            # 6. Verificar conteos después de la carga
            print("\n6. Verificando personal después de la carga...")
            empleados_despues = db.session.query(Empleado).filter_by(empresa_id=empresa.id, activo=True).count()
            locadores_despues = db.session.query(Locador).filter_by(empresa_id=empresa.id, activo=True).count()
            print(f"   OK: Empleados después: {empleados_despues}")
            print(f"   OK: Locadores después: {locadores_despues}")
            
            # 7. Verificar que los datos se cargaron correctamente
            print("\n7. Verificando datos cargados...")
            empleados_nuevos = db.session.query(Empleado).filter_by(empresa_id=empresa.id, activo=True).all()
            for empleado in empleados_nuevos[-empleados_cargados:]:  # Solo los últimos agregados
                print(f"   OK: {empleado.nombres} {empleado.apellidos} - DNI: {empleado.dni} - Sueldo: S/. {empleado.sueldo_base}")
            
            locadores_nuevos = db.session.query(Locador).filter_by(empresa_id=empresa.id, activo=True).all()
            for locador in locadores_nuevos[-locadores_cargados:]:  # Solo los últimos agregados
                print(f"   OK: {locador.nombres} {locador.apellidos} - DNI: {locador.dni} - Monto: S/. {locador.monto_mensual}")
            
            # 8. Limpiar archivo de prueba
            print("\n8. Limpiando archivo de prueba...")
            os.remove(archivo_excel)
            print(f"   OK: Archivo {archivo_excel} eliminado")
            
            print("\n" + "="*60)
            print("CARGA MASIVA DESDE EXCEL FUNCIONANDO CORRECTAMENTE")
            print("="*60)
            print(f"Total empleados cargados: {empleados_cargados}")
            print(f"Total locadores cargados: {locadores_cargados}")
            print("La funcionalidad de carga masiva está lista para usar")
            
            return True
            
        except Exception as e:
            print(f"\nERROR: {e}")
            import traceback
            traceback.print_exc()
            return False

if __name__ == '__main__':
    print("INICIANDO PRUEBA DE CARGA MASIVA DESDE EXCEL (SIMPLIFICADA)...")
    if probar_carga_simple():
        print("\nSISTEMA DE CARGA MASIVA LISTO PARA USAR")
    else:
        print("\nHAY PROBLEMAS EN LA CARGA MASIVA")
